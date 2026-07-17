#!/usr/bin/env python3
"""Download and normalise the freshest national BES regional workbook.

The Istat intermediate BES release carries regional observations that are often
one year fresher than BES dei Territori.  This importer keeps the same 12-column
dataset contract and manifest contract already consumed by ``app.bes_data``.
Only total-sex rows for the 20 regions are retained.
"""

from __future__ import annotations

import argparse
import csv
import html
import io
import os
import re
import sys
import tempfile
import urllib.parse
import urllib.request
import zipfile
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.bes_national_sources import category_for, direction_for
from scripts.update_data import OUTPUT_COLUMNS, REGION_NAME_MAP, SUPPORTED_REGIONS


SOURCE_PAGE = (
    "https://www.istat.it/statistiche-per-temi/focus/benessere-e-sostenibilita/"
    "la-misurazione-del-benessere-bes/gli-indicatori-del-bes/"
)
FALLBACK_ARTIFACT = "https://www.istat.it/wp-content/uploads/2026/05/APPENDICE-STATISTICA-2.zip"
OUTPUT = PROJECT_ROOT / "app" / "static" / "data" / "Assoluti_BES_Regione.csv"
MANIFEST = PROJECT_ROOT / "app" / "static" / "data" / "bes_regione_manifest.csv"
WORKBOOK_BASENAME = "indicatori_regione_sesso.xlsx"
MANIFEST_COLUMNS = [
    "id", "name", "domain", "domain_name", "proposed_category",
    "proposed_direction", "unit", "year_min", "year_max", "n_region",
    "coverage", "n_region_latest", "coverage_latest", "source_dataflow",
]

REGION_ALIASES = {
    **REGION_NAME_MAP,
    "Trentino-Alto Adige/Südtirol": "Trentino Alto Adige",
    "Friuli-Venezia Giulia": "Friuli-Venezia Giulia",
}


def _request(url):
    return urllib.request.Request(
        url,
        headers={"User-Agent": "diset-viz-data-updater/1.0 (+https://divarioitalia.it)"},
    )


def discover_artifact_url(page_url=SOURCE_PAGE):
    """Return the first current BES statistical-appendix ZIP linked by Istat."""
    with urllib.request.urlopen(_request(page_url), timeout=60) as response:
        page = response.read().decode("utf-8", errors="replace")
    hrefs = re.findall(r'href=["\']([^"\']+)["\']', page, flags=re.IGNORECASE)
    for href in hrefs:
        decoded = html.unescape(href)
        name = decoded.rsplit("/", 1)[-1].lower()
        if "appendice" in name and "statistica" in name and name.endswith(".zip"):
            return urllib.parse.urljoin(page_url, decoded)
    return FALLBACK_ARTIFACT


def download_artifact(url):
    with urllib.request.urlopen(_request(url), timeout=120) as response:
        return response.read(), {
            "url": response.geturl(),
            "etag": response.headers.get("ETag", ""),
            "last_modified": response.headers.get("Last-Modified", ""),
        }


def _workbook_bytes(archive_bytes):
    with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
        candidates = [
            name for name in archive.namelist()
            if Path(name).name.lower() == WORKBOOK_BASENAME
        ]
        if not candidates:
            candidates = [
                name for name in archive.namelist()
                if name.lower().endswith(".xlsx")
                and "regione" in Path(name).name.lower()
                and "sesso" in Path(name).name.lower()
            ]
        if len(candidates) != 1:
            raise RuntimeError(
                f"Expected one {WORKBOOK_BASENAME} in BES ZIP, found {len(candidates)}"
            )
        return archive.read(candidates[0])


def _number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace("\xa0", "")
    if not text or text in {"-", "..", "...", ":", "n.d.", "nd"}:
        return None
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _italian_number(value):
    if isinstance(value, int) or (isinstance(value, float) and value.is_integer()):
        return str(int(value))
    return format(float(value), ".12g").replace(".", ",")


def _region_name(raw):
    name = " ".join(str(raw or "").split())
    return REGION_ALIASES.get(name, name)


def parse_archive(archive_bytes):
    workbook = load_workbook(
        io.BytesIO(_workbook_bytes(archive_bytes)), read_only=True, data_only=True
    )
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    positions = {str(value).strip(): index for index, value in enumerate(header)}
    required = {"DOMINIO", "CODICE", "INDICATORE", "SESSO", "TERRITORIO", "UNITA_MISURA", "FONTE"}
    missing = required - set(positions)
    if missing:
        raise RuntimeError(f"Missing BES workbook columns: {sorted(missing)}")
    year_columns = [(index, int(value)) for index, value in enumerate(header) if isinstance(value, int)]
    if not year_columns:
        raise RuntimeError("No year columns found in BES workbook")

    dataset = []
    observations = defaultdict(lambda: {"years": set(), "regions": set(), "by_year": defaultdict(set)})
    metadata = {}
    for values in rows:
        if str(values[positions["SESSO"]] or "").strip() != "Totale":
            continue
        territory = _region_name(values[positions["TERRITORIO"]])
        if territory not in SUPPORTED_REGIONS:
            continue
        indicator_id = str(values[positions["CODICE"]] or "").strip()
        if not indicator_id:
            continue
        domain = str(values[positions["DOMINIO"]] or "").strip()
        name = " ".join(str(values[positions["INDICATORE"]] or "").split())
        unit = " ".join(str(values[positions["UNITA_MISURA"]] or "").split())
        source = " ".join(str(values[positions["FONTE"]] or "").split())
        metadata[indicator_id] = {"domain": domain, "name": name, "unit": unit}
        for index, year in year_columns:
            value = _number(values[index])
            if value is None:
                continue
            dataset.append({
                "idIndicatore": indicator_id,
                "Territorio": territory,
                "Tema": domain,
                "Indicatore": name,
                "UDM": unit,
                "Fonte": source or "Istat",
                "Archivio": "Benessere equo e sostenibile, aggiornamento intermedio 2026",
                "Anno": str(year),
                "Livello/Variazione": "Livello",
                "Dato": _italian_number(value),
                "Benchmark": "",
                "Area": "Regione",
            })
            info = observations[indicator_id]
            info["years"].add(year)
            info["regions"].add(territory)
            info["by_year"][year].add(territory)

    if not dataset:
        raise RuntimeError("No regional BES observations parsed")
    dataset.sort(key=lambda row: (row["idIndicatore"], row["Territorio"], int(row["Anno"])))

    manifest = []
    for indicator_id, info in observations.items():
        meta = metadata[indicator_id]
        year_max = max(info["years"])
        latest = len(info["by_year"][year_max])
        manifest.append({
            "id": indicator_id,
            "name": meta["name"],
            "domain": meta["domain"],
            "domain_name": meta["domain"],
            "proposed_category": category_for(meta["domain"]) or "",
            "proposed_direction": direction_for(indicator_id, meta["name"]),
            "unit": meta["unit"],
            "year_min": min(info["years"]),
            "year_max": year_max,
            "n_region": len(info["regions"]),
            "coverage": round(len(info["regions"]) / 20, 4),
            "n_region_latest": latest,
            "coverage_latest": round(latest / 20, 4),
            "source_dataflow": "BES nazionale, appendice statistica regionale",
        })
    manifest.sort(key=lambda row: (row["domain"], row["id"]))
    return dataset, manifest


def _write_atomic(rows, columns, output_path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w", encoding="utf-8", newline="", delete=False, dir=output_path.parent
    ) as tmp:
        writer = csv.DictWriter(tmp, fieldnames=columns, delimiter=";", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
        temp_name = tmp.name
    os.replace(temp_name, output_path)
    os.chmod(output_path, 0o644)


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, help="Use a cached BES ZIP instead of the network")
    parser.add_argument("--source-page", default=SOURCE_PAGE)
    parser.add_argument("--artifact-url", default="")
    parser.add_argument("--output", type=Path, default=OUTPUT)
    parser.add_argument("--manifest", type=Path, default=MANIFEST)
    return parser.parse_args()


def main():
    args = parse_args()
    if args.input:
        archive_bytes = args.input.read_bytes()
        artifact_url = str(args.input)
    else:
        artifact_url = args.artifact_url or discover_artifact_url(args.source_page)
        archive_bytes, _ = download_artifact(artifact_url)
    dataset, manifest = parse_archive(archive_bytes)
    _write_atomic(dataset, OUTPUT_COLUMNS, args.output)
    _write_atomic(manifest, MANIFEST_COLUMNS, args.manifest)
    current = sum(1 for row in manifest if int(row["year_max"]) >= 2025)
    scoreable_current = sum(
        1 for row in manifest
        if int(row["year_max"]) >= 2025
        and row["proposed_direction"] != "contextual"
        and float(row["coverage_latest"]) >= 0.8
        and row["proposed_category"]
    )
    print(
        f"BES source: {artifact_url}\n"
        f"Wrote {len(dataset)} rows and {len(manifest)} indicators; "
        f"{current} reach 2025, {scoreable_current} are current and scoreable."
    )


if __name__ == "__main__":
    main()
