#!/usr/bin/env python3
"""Read-only audit: do the already-downloaded Istat sources hold more
dimension data (sex, secondary theme) than the converters currently read?

Downloads the same two archives scripts/update_data.py and
scripts/update_bes_regions.py already fetch, and reports what their
converters leave on the floor: unread columns, rows dropped by the
territory filter, and (for the historical catalogue) a cross-check of the
title-suffix heuristic against Istat's own gender-breakdown tag.

Written for docs/FAMIGLIE_INDICATORI.md (Settimana 3-4, passo a), after
Nello asked whether the sources already acquired had more of this than we
were using. Does not touch any file under app/static/data: report only.
"""

from __future__ import annotations

import csv
import io
import re
import sys
import tempfile
import urllib.request
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.update_data import REGION_NAME_MAP, SOURCE_URL, SUPPORTED_REGIONS  # noqa: E402
from scripts.update_bes_regions import (  # noqa: E402
    REGION_ALIASES,
    discover_artifact_url,
)

# Columns scripts/update_data.py:convert_row actually reads.
CATALOG_COLUMNS_USED = {
    "COD_INDICATORE", "TITOLO", "SOTTOTITOLO", "ANNO_RIFERIMENTO",
    "UNITA_MISURA", "VALORE", "DESCRIZIONE_RIPARTIZIONE",
    "DESCRIZIONE_TEMA1", "OC_TEMA_SINTETICO", " 1° OBIETTIVO",
}

GENDER_TAG = "Asse VII - Articolazione di genere."

GENDER_SUFFIX_RE = re.compile(
    r"""(?:\s*\(\s*(?P<g1>maschi|femmine|totale)\s*\)\s*$
          |\s*[-,]\s*(?P<g2>maschi|femmine|totale)\s*$
          |\s+(?P<g3>maschi|femmine|totale)\s*$)""",
    re.IGNORECASE | re.VERBOSE,
)


def _download(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "diset-viz-audit/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def audit_historical_catalog() -> None:
    print("=" * 70)
    print("CATALOGO STORICO (Assoluti_Regione.csv, fonte BDTPS)")
    print("=" * 70)
    archive = zipfile.ZipFile(io.BytesIO(_download(SOURCE_URL)))
    csv_names = [n for n in archive.namelist() if n.lower().endswith(".csv")]
    with archive.open(csv_names[0]) as raw:
        text = io.TextIOWrapper(raw, encoding="utf-8-sig", newline="")
        reader = csv.DictReader(text, delimiter=";")
        header = reader.fieldnames or []
        unused_columns = [c for c in header if c not in CATALOG_COLUMNS_USED]
        print(f"\nColonne nel CSV Istat: {len(header)}, lette da convert_row: "
              f"{len(CATALOG_COLUMNS_USED)}, mai lette: {len(unused_columns)}")
        print(f"  {unused_columns}")

        territory_counts = Counter()
        secondary_themes = set()
        gender_tagged_ids = set()
        title_by_id = {}
        n_rows = 0
        for row in reader:
            n_rows += 1
            territory_counts[row["DESCRIZIONE_RIPARTIZIONE"]] += 1
            if row["DESCRIZIONE_TEMA2"].strip():
                secondary_themes.add(row["DESCRIZIONE_TEMA2"].strip())
            territory = REGION_NAME_MAP.get(
                row["DESCRIZIONE_RIPARTIZIONE"], row["DESCRIZIONE_RIPARTIZIONE"]
            )
            if territory in SUPPORTED_REGIONS:
                iid = row["COD_INDICATORE"].lstrip("0") or "0"
                title_by_id[iid] = row["TITOLO"].strip()
                if row["DESCRIZIONE_ASSE_QCS"].strip() == GENDER_TAG:
                    gender_tagged_ids.add(iid)

    kept_rows = sum(v for k, v in territory_counts.items()
                     if REGION_NAME_MAP.get(k, k) in SUPPORTED_REGIONS)
    print(f"\nRighe totali: {n_rows}, tenute (20 regioni): {kept_rows}, "
          f"scartate per territorio: {n_rows - kept_rows}")
    for k, v in sorted(territory_counts.items(), key=lambda x: -x[1]):
        if REGION_NAME_MAP.get(k, k) not in SUPPORTED_REGIONS:
            print(f"  scartato: {k!r}: {v} righe")

    print(f"\nDESCRIZIONE_TEMA2 (seconda classificazione, mai letta): "
          f"{len(secondary_themes)} valori -> {sorted(secondary_themes)}")
    print(f"\nIndicatori taggati {GENDER_TAG!r} (ufficiale Istat, mai letto): "
          f"{len(gender_tagged_ids)} su {len(title_by_id)}")

    regex_gender_ids = set()
    for iid, title in title_by_id.items():
        m = GENDER_SUFFIX_RE.search(title)
        if m:
            g = (m.group("g1") or m.group("g2") or m.group("g3")).lower()
            if g in ("maschi", "femmine"):
                regex_gender_ids.add(iid)
    regex_only = regex_gender_ids - gender_tagged_ids
    tag_only = gender_tagged_ids - regex_gender_ids
    outcome = "combaciano" if not regex_only and not tag_only else "DISCREPANZE"
    print(f"Regex sui titoli vs tag ufficiale (solo appartenenza, non il "
          f"valore maschi/femmine: vedi nota sotto): {len(regex_only)} solo "
          f"regex, {len(tag_only)} solo tag -> {outcome}")
    print("Nota: DESCRIZIONE_ASSE_QCS ha lo stesso valore per tutti gli "
          "indicatori taggati: dice che l'id fa parte di una scomposizione "
          "di genere, non se e' maschi o femmine ne' quale totale gli "
          "corrisponde. Non sostituisce il parser sul titolo per costruire "
          "le famiglie, lo conferma come insieme di appartenenza.")


def audit_bes() -> None:
    print()
    print("=" * 70)
    print("BES (Assoluti_BES_Regione.csv, fonte Appendice Statistica)")
    print("=" * 70)
    artifact_url = discover_artifact_url()
    print(f"\nURL scoperto da discover_artifact_url() (stesso path del "
          f"convertitore): {artifact_url}")
    with zipfile.ZipFile(io.BytesIO(_download(artifact_url))) as archive:
        xlsx_names = [n for n in archive.namelist() if n.lower().endswith(".xlsx")]
        print(f"\nFile Excel nello ZIP: {len(xlsx_names)}")
        for n in xlsx_names:
            used = "usato" if Path(n).name.lower() == "indicatori_regione_sesso.xlsx" else "mai aperto"
            print(f"  {n} [{used}]")

        target = next(n for n in xlsx_names if Path(n).name.lower() == "indicatori_regione_sesso.xlsx")
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            tmp.write(archive.read(target))
            tmp.flush()
            workbook = load_workbook(tmp.name, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header = next(rows)
            pos = {v: i for i, v in enumerate(header)}

            sex_row_counts = Counter()
            regions_by_indicator_sex = defaultdict(lambda: defaultdict(set))
            territories_seen = set()
            for values in rows:
                sex = str(values[pos["SESSO"]] or "").strip()
                indicator_id = str(values[pos["CODICE"]] or "").strip()
                raw_territory = " ".join(str(values[pos["TERRITORIO"]] or "").split())
                territories_seen.add(raw_territory)
                sex_row_counts[sex] += 1
                territory = REGION_ALIASES.get(raw_territory, raw_territory)
                if territory in SUPPORTED_REGIONS:
                    regions_by_indicator_sex[indicator_id][sex].add(territory)
            workbook.close()

    print(f"\nRighe per SESSO (tutti i territori del file): {dict(sex_row_counts)}")

    complete_triplets = set()
    for indicator_id, by_sex in regions_by_indicator_sex.items():
        if all(
            by_sex.get(sex, set()) == SUPPORTED_REGIONS
            for sex in ("Totale", "Maschi", "Femmine")
        ):
            complete_triplets.add(indicator_id)
    totale_only_ids = {
        indicator_id for indicator_id, by_sex in regions_by_indicator_sex.items()
        if by_sex.get("Totale", set()) == SUPPORTED_REGIONS
        and not by_sex.get("Maschi") and not by_sex.get("Femmine")
    }
    any_gender_ids = {
        indicator_id for indicator_id, by_sex in regions_by_indicator_sex.items()
        if by_sex.get("Maschi") or by_sex.get("Femmine")
    }
    print(f"Indicatori con Maschi e/o Femmine in almeno una regione: {len(any_gender_ids)}")
    print(f"Indicatori con tripletta Totale+Maschi+Femmine COMPLETA su tutte "
          f"e 20 le regioni: {len(complete_triplets)}")
    print(f"Indicatori Totale-only su tutte le 20 regioni, nessun "
          f"Maschi/Femmine in nessuna regione: {len(totale_only_ids)}")
    print(f"Righe scartate dal filtro 'solo Totale' (tutti i territori del "
          f"file, non solo le 20 regioni): "
          f"{sex_row_counts.get('Maschi', 0) + sex_row_counts.get('Femmine', 0)}")
    print(f"Territori nel file oltre le 20 regioni (scartati, dopo alias): "
          f"{sorted(t for t in territories_seen if REGION_ALIASES.get(t, t) not in SUPPORTED_REGIONS)}")


if __name__ == "__main__":
    audit_historical_catalog()
    audit_bes()
