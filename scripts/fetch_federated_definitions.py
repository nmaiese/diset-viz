#!/usr/bin/env python3
"""Build the committed definition archive for every non-native source family.

The native territorial sheet remains owned by ``scripts/fetch_definitions``.
This companion fetcher covers the holes that sheet cannot contain:

- the discontinued territorial indicator and the locally assigned 9xx ids;
- BES, from ``Metadati.xlsx`` in Istat's current statistical appendix;
- Multiscopo and demographic series, from the exact Istat SDMX codelist values
  pinned by the committed source configurations;
- the two Eurostat R&D series, from Eurostat's current R&D metadata.

Web material is treated as source data only.  Definitions are written with the
source URL and the precise workbook/codelist reference that supports them.

    .venv/bin/python scripts/fetch_federated_definitions.py
    .venv/bin/python scripts/fetch_federated_definitions.py --dry-run
    .venv/bin/python scripts/fetch_federated_definitions.py \
        --bes-zip APPENDICE-STATISTICA.zip --cache-dir data/istat_cache

The Istat SDMX endpoint allows at most five requests per minute.  The shared
client spaces cache misses by 16 seconds; a warm cache makes reruns immediate.
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from app import sources  # noqa: E402
from scripts import (  # noqa: E402
    eurostat_source,
    istat_regional_source,
    istat_sdmx,
    multiscopo_sources,
    prose_lint,
    update_bes_regions,
)
from scripts.fetch_definitions import (  # noqa: E402
    FEDERATED_OUTPUT_PATH,
    OUTPUT_PATH as TERRITORIAL_DEFINITIONS,
    _clean,
    _load_csv,
    write_atomic,
)

BES_ARTIFACT_URL = update_bes_regions.FALLBACK_ARTIFACT
BES_TERRITORIES_ARTIFACT_URL = (
    "https://www.istat.it/wp-content/uploads/2024/07/"
    "Bes_dei_territori_edizione_2024_18122024_IT_EC.zip"
)
EUROSTAT_METADATA_URL = "https://ec.europa.eu/eurostat/cache/metadata/en/rd_esms.htm"
ISTAT_SDMX_URL = "https://esploradati.istat.it/SDMXWS/rest"
CURATION_PATH = PROJECT_ROOT / "data" / "discovery" / "curation.csv"
ATLAS_PATH = PROJECT_ROOT / "app" / "static" / "data" / "Assoluti_Regione.csv"

FEDERATED_COLUMNS = [
    "id",
    "indicatore",
    "definizione",
    "fonti",
    "source_url",
    "source_reference",
]

# The source itself no longer lists this discontinued table.  This is Istat's
# historical A.08 definition, including the note that prevents "soglia" from
# being read as one fixed euro amount across years.
DISCONTINUED = {
    "76": {
        "indicatore": "Indice di povertà regionale (famiglie)",
        "definizione": (
            "Famiglie con spesa media mensile per consumi pari o inferiore alla "
            "soglia di povertà sul totale delle famiglie residenti, per cento. "
            "La soglia è calcolata rispetto al livello medio dei consumi di ogni anno."
        ),
        "fonti": "Istat, Indagine sui consumi delle famiglie",
        "source_url": "https://www.istat.it/wp-content/uploads/2011/01/Documentazione.pdf",
        "source_reference": "Indicatore A.08, tavola territoriale dismessa",
    }
}

EUROSTAT_DEFINITIONS = {
    "rd_e_gerdreg": (
        "Spesa interna lorda per ricerca e sviluppo sostenuta da tutti i settori "
        "di esecuzione, espressa in percentuale del prodotto interno lordo regionale."
    ),
    "rd_p_persreg": (
        "Personale direttamente impegnato in ricerca e sviluppo o che presta servizi "
        "diretti alle attività di ricerca e sviluppo, misurato in equivalenti a tempo "
        "pieno ed espresso in percentuale della popolazione attiva."
    ),
}


def _record(key, title, definition, source, url, reference):
    return {
        "id": key,
        "indicatore": _typographic(_clean(title)),
        "definizione": _finish(_typographic(_clean(definition))),
        "fonti": _typographic(_clean(source)),
        "source_url": _clean(url),
        "source_reference": _clean(reference),
    }


def _finish(text):
    return text if not text or text.endswith((".", "?", "!")) else f"{text}."


def _typographic(text):
    return (
        text.replace("Età", "Età")
        .replace("età", "età")
        .replace("più", "più")
        .replace("può", "può")
        .replace("già", "già")
        .replace("mortalità", "mortalità")
        .replace("ospitalità", "ospitalità")
        .replace("qualità", "qualità")
    )


def _read_rows(path):
    with Path(path).open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=";"))


def _bes_metadata(archive_bytes, source_url=BES_ARTIFACT_URL):
    with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
        candidates = [
            name for name in archive.namelist()
            if Path(name).name.lower() == "metadati.xlsx"
        ]
        if len(candidates) != 1:
            raise RuntimeError(
                f"Expected one Metadati.xlsx in the BES archive, found {len(candidates)}"
            )
        workbook = load_workbook(
            io.BytesIO(archive.read(candidates[0])), read_only=True, data_only=True
        )
    sheet = workbook["Metadati"] if "Metadati" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    positions = {str(value or "").strip().lower(): index for index, value in enumerate(header)}
    required = {"codice", "indicatore", "definizione", "fonte"}
    missing = required - set(positions)
    if missing:
        raise RuntimeError(f"Missing BES metadata columns: {sorted(missing)}")

    out = []
    for values in rows:
        code = _clean(values[positions["codice"]])
        definition = _clean(values[positions["definizione"]])
        if not code or not definition:
            continue
        out.append(_record(
            f"bes:{code}",
            values[positions["indicatore"]],
            definition,
            values[positions["fonte"]],
            source_url,
            f"Metadati.xlsx, codice {code}",
        ))
    return out


def _bes_territories_metadata(archive_bytes):
    """BES dei Territori definitions absent from the national BES workbook."""
    with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
        candidates = [
            name for name in archive.namelist()
            if Path(name).name.lower().startswith("metadati_it")
            and name.lower().endswith(".xlsx")
        ]
        if len(candidates) != 1:
            raise RuntimeError(
                "Expected one Italian metadata workbook in the BES dei Territori "
                f"archive, found {len(candidates)}"
            )
        workbook = load_workbook(
            io.BytesIO(archive.read(candidates[0])), read_only=True, data_only=True
        )
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    positions = {str(value or "").replace("\n", " ").strip().lower(): index
                 for index, value in enumerate(header)}
    required = {"cod. indicatore", "indicatore", "definizione", "fonte"}
    missing = required - set(positions)
    if missing:
        raise RuntimeError(f"Missing BES dei Territori metadata columns: {sorted(missing)}")

    out = []
    empty = 0
    for values in rows:
        code = _clean(values[positions["cod. indicatore"]])
        if not code:
            empty += 1
            if empty >= 50:
                break
            continue
        empty = 0
        definition = _clean(values[positions["definizione"]])
        if not definition:
            continue
        source_parts = [values[positions["fonte"]]]
        if positions["fonte"] + 1 < len(values):
            source_parts.append(values[positions["fonte"] + 1])
        out.append(_record(
            f"bes:{code}",
            values[positions["indicatore"]],
            definition,
            ", ".join(_clean(part) for part in source_parts if _clean(part)),
            BES_TERRITORIES_ARTIFACT_URL,
            f"Metadati_IT_ed.2024.xlsx, codice {code}",
        ))
    return out


def _territorial_local_definitions():
    existing = set(_load_csv(TERRITORIAL_DEFINITIONS))
    article_ids = {
        key for key in prose_lint.load_texts()
        if key.isdigit() and (key == "76" or key.startswith("9"))
    }
    first = {}
    for row in _read_rows(ATLAS_PATH):
        key = row["idIndicatore"]
        if key in article_ids and key not in first:
            first[key] = row

    out = [
        _record(
            key,
            record["indicatore"],
            record["definizione"],
            record["fonti"],
            record["source_url"],
            record["source_reference"],
        )
        for key, record in DISCONTINUED.items()
        if key in article_ids and key not in existing
    ]
    for key in sorted(article_ids - existing - set(DISCONTINUED), key=int):
        row = first.get(key)
        if row is None:
            raise RuntimeError(f"No territorial source row for {key}")
        provenance = sources.territorial_source(key)
        out.append(_record(
            key,
            row["Indicatore"],
            _typographic(row["Archivio"]),
            provenance["label"],
            provenance["url"],
            f"Assoluti_Regione.csv, idIndicatore {key}",
        ))
    return out


def _codelist_id(urn):
    match = re.search(r"Codelist=[^:]+:([^()]+)", urn or "")
    return match.group(1) if match else ""


def _sdmx_metadata(client, specs):
    """Official labels for every pinned dimension of each Multiscopo series."""
    flows = {flow["id"]: flow for flow in client.dataflows()}
    structures = {}
    codelists = {}
    result = {}
    for spec in specs:
        flow_id = spec["flow_id"]
        flow = flows.get(flow_id)
        if flow is None:
            raise RuntimeError(f"Istat dataflow disappeared: {flow_id}")
        dsd_match = re.search(r"DataStructure=[^:]+:([^()]+)", flow["dsd"])
        if not dsd_match:
            raise RuntimeError(f"No DSD reference for {flow_id}")
        dsd_id = dsd_match.group(1)
        if dsd_id not in structures:
            structures[dsd_id] = client.datastructure(dsd_id)
        structure = structures[dsd_id]
        dimensions = {dim["id"]: _codelist_id(dim["codelist"])
                      for dim in structure["dimensions"]}
        labels = {}
        references = []
        for dimension, code in spec["filters"].items():
            codelist_id = dimensions.get(dimension)
            if not codelist_id:
                raise RuntimeError(f"{flow_id}: no codelist for {dimension}")
            if codelist_id not in codelists:
                codelists[codelist_id] = client.codelist(codelist_id)
            codelist = codelists[codelist_id]
            item = codelist["codes"].get(code)
            if not item or not item["name"]:
                raise RuntimeError(
                    f"{flow_id}: {codelist_id} has no label for {dimension}={code}"
                )
            labels[dimension] = item["name"]
            references.append(f"{codelist_id}:{code}={item['name']}")
        result[spec["id"]] = {
            "flow_name": flow["name"],
            "labels": labels,
            "reference": f"{flow_id}; " + "; ".join(references),
        }
    return result


def _multiscopo_definition(spec, metadata):
    labels = metadata["labels"]
    data_type = labels.get("DATA_TYPE", "")
    measure = labels.get("MEASURE", "")
    key = spec["id"]

    if key.startswith("MULTI_SATLIFE_") or key.startswith("MULTI_ELETTR_") \
            or key.startswith("MULTI_BMI_"):
        return f"{data_type.capitalize()}, {measure}."
    if key.startswith("MULTI_ICT_"):
        return f"Famiglie con {data_type}, valori percentuali."
    if key in {"MULTI_ABIT_PROBLEMI", "MULTI_ABIT_INFISSI_FATISCENTI", "MULTI_ABIT_UMIDITA"}:
        problem = labels["PROBLEM_WITH_ACCOMM"]
        return (
            f"Famiglie con problemi di {problem} nell'abitazione, per 100 famiglie "
            "con le stesse caratteristiche."
        )
    if key in {"MULTI_ABIT_AFFITTO", "MULTI_ABIT_PROPRIETA"}:
        return (
            "Famiglie per titolo di godimento dell'abitazione: "
            f"{labels['TENURE_STATUS']}, valori percentuali."
        )
    if key.startswith("MULTI_REDD_FONTE_"):
        return (
            f"{data_type.capitalize()}: {labels['FAM_MAIN_INCOME_SOURCE']}, {measure}."
        )
    if key in {"MULTI_REDD_MEDIANO", "MULTI_REDD_MEDIO"}:
        return f"{data_type.capitalize()}, {labels['IMPUTED_RENTS']}."
    if key.startswith("MULTI_DISAGIO_") and "ECON_SITUATION_PERCEIVED" in labels:
        return f"{data_type.capitalize()}: {labels['ECON_SITUATION_PERCEIVED']}."
    if key.startswith("MULTI_ZONA_"):
        return (
            f"Famiglie che dichiarano {labels['PROBL_DWELLING_PLACE']} nella zona "
            "di residenza, valori per cento."
        )
    if key == "MULTI_SPESA_GINI":
        return "Indice di concentrazione di Gini della spesa per consumi delle famiglie."
    if key == "MULTI_SPESA_INTERQUINTILE":
        return (
            "Rapporto tra la spesa per consumi del quinto di famiglie con la spesa "
            "più alta e quella del quinto con la spesa più bassa."
        )
    return data_type


def _multiscopo_definitions(client):
    specs = multiscopo_sources.MULTISCOPO_INDICATORS
    metadata = _sdmx_metadata(client, specs)
    return [
        _record(
            f"multiscopo:{spec['id']}",
            spec["name"],
            _multiscopo_definition(spec, metadata[spec["id"]]),
            "Istat, Indagine Multiscopo sulle famiglie",
            f"{ISTAT_SDMX_URL}/data/{spec['flow_id']}",
            metadata[spec["id"]]["reference"],
        )
        for spec in specs
    ]


def _curated_descriptions():
    return {
        row["target_indicator_id"]: row
        for row in _read_rows(CURATION_PATH)
        if row.get("description")
    }


def _first_sentence(text):
    return re.split(r"(?<=\.)\s+", _typographic(text.strip()), maxsplit=1)[0]


def _demographic_definitions(client):
    codelist = client.codelist("CL_TIPO_DATO15")["codes"]
    curated = _curated_descriptions()
    out = []
    for raw_id, spec in sorted(istat_regional_source.ISTAT_SERIES.items()):
        key = f"dem:{raw_id}"
        if key not in curated:
            continue
        official = codelist.get(raw_id)
        if not official or not official["name"]:
            raise RuntimeError(f"CL_TIPO_DATO15 has no label for {raw_id}")
        out.append(_record(
            key,
            spec["name"],
            _first_sentence(curated[key]["description"]),
            "Istat, Indicatori demografici",
            ISTAT_SDMX_URL,
            f"CL_TIPO_DATO15:{raw_id}={official['name']}",
        ))
    return out


def _eurostat_definitions():
    return [
        _record(
            f"eur:{raw_id}",
            eurostat_source.EUROSTAT_SERIES[raw_id]["name"],
            definition,
            "Eurostat, Research and development statistics",
            EUROSTAT_METADATA_URL,
            (
                f"dataset {eurostat_source.EUROSTAT_SERIES[raw_id]['dataset']}; "
                + "; ".join(
                    f"{name}={value}"
                    for name, value in eurostat_source.EUROSTAT_SERIES[raw_id]["params"].items()
                )
            ),
        )
        for raw_id, definition in EUROSTAT_DEFINITIONS.items()
    ]


def build_records(
    bes_archive,
    bes_territories_archive,
    client,
    bes_source_url=BES_ARTIFACT_URL,
):
    national_bes = _bes_metadata(bes_archive, source_url=bes_source_url)
    national_ids = {row["id"] for row in national_bes}
    local_bes = [
        row for row in _bes_territories_metadata(bes_territories_archive)
        if row["id"] not in national_ids
    ]
    records = (
        _territorial_local_definitions()
        + national_bes
        + local_bes
        + _multiscopo_definitions(client)
        + _demographic_definitions(client)
        + _eurostat_definitions()
    )
    by_id = {}
    for record in records:
        if record["id"] in by_id:
            raise RuntimeError(f"Duplicate definition id: {record['id']}")
        if not record["definizione"] or not record["source_url"]:
            raise RuntimeError(f"Incomplete definition: {record['id']}")
        by_id[record["id"]] = record
    return [by_id[key] for key in sorted(by_id)]


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("--bes-zip", type=Path)
    parser.add_argument("--best-zip", type=Path,
                        help="BES dei Territori archive with province-only metadata")
    parser.add_argument("--cache-dir", type=Path,
                        default=PROJECT_ROOT / "data" / "istat_cache")
    parser.add_argument("--output", type=Path, default=FEDERATED_OUTPUT_PATH)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    if args.bes_zip:
        bes_archive = args.bes_zip.read_bytes()
        bes_source_url = BES_ARTIFACT_URL
    else:
        bes_source_url = update_bes_regions.discover_artifact_url()
        bes_archive, _ = update_bes_regions.download_artifact(bes_source_url)
    if args.best_zip:
        bes_territories_archive = args.best_zip.read_bytes()
    else:
        bes_territories_archive, _ = update_bes_regions.download_artifact(
            BES_TERRITORIES_ARTIFACT_URL
        )

    client = istat_sdmx.SdmxClient(args.cache_dir, min_interval=16)
    records = build_records(
        bes_archive,
        bes_territories_archive,
        client,
        bes_source_url=bes_source_url,
    )
    previous = _load_csv(args.output)
    changed = [
        row["id"] for row in records
        if previous.get(row["id"]) != row
    ]
    gone = sorted(set(previous) - {row["id"] for row in records})
    print(f"{len(records)} definizioni federate")
    print(f"  nuove o cambiate  {len(changed)}")
    print(f"  sparite           {len(gone)}")
    print(f"  richieste SDMX    {client.request_count}")
    if args.dry_run:
        return 0
    write_atomic(records, args.output, columns=FEDERATED_COLUMNS)
    print(f"scritto {args.output.relative_to(PROJECT_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
